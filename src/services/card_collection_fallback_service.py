from __future__ import annotations

from dataclasses import replace
from datetime import date, datetime
from pathlib import Path
import re
from typing import Any

from src.models.schemas import NormalizedDataset

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


class CardCollectionFallbackService:
    CARRY_FORWARD_HEADERS = {"大区", "地区", "原销售大区", "纸袋分类", "滔搏纸袋分类", "型号", "规格", "盘点月", "日期 (月)", "日期", "月份"}
    MONTH_ALIGNED_CARD_IDS = {
        "xe5da9d423db44bbe96028ad",
        "a597c4441b7414c93a7c502d",
        "l6e08fdcc7fef45ccaa31d1b",
    }
    PERIOD_HEADERS = {"日期 (月)", "盘点日 (月)", "日期", "月份", "月", "统计月份"}
    TOTAL_TOKENS = ("总计", "合计", "小计")

    def __init__(self, collection_dir: Path, logger: Any, sheet_overrides: dict[str, str] | None = None) -> None:
        self.collection_dir = collection_dir
        self.logger = logger
        self.sheet_overrides = sheet_overrides or {}

    def apply(self, datasets: list[NormalizedDataset], report_month: str | None = None) -> list[NormalizedDataset]:
        return [self._apply_to_dataset(dataset, report_month=report_month) for dataset in datasets]

    def _apply_to_dataset(self, dataset: NormalizedDataset, report_month: str | None = None) -> NormalizedDataset:
        application_errors = dataset.summary.get("application_errors", [])
        fallback_path = self.collection_dir / f"{dataset.card_id}.xlsx"
        if not fallback_path.exists() or openpyxl is None:
            return dataset

        workbook = openpyxl.load_workbook(fallback_path, read_only=True, data_only=True)
        worksheet = self._resolve_sheet(workbook, dataset.card_id)
        if worksheet is None:
            return dataset

        rows = self._extract_rows(worksheet)
        if not rows:
            return dataset

        rows = self._normalize_rows(dataset.card_id, rows, report_month=report_month)
        aligned_rows = self._month_aligned_rows(dataset, rows, report_month)
        if aligned_rows is not None:
            fallback_info = {
                "source_type": "card_collection_month_alignment",
                "fallback_path": str(fallback_path),
                "sheet_name": worksheet.title,
                "report_month": report_month,
                "replaced_api_rows": bool(dataset.rows),
                "reason": "API result did not match the exported card collection SQL/month scope.",
            }
            summary = {
                "row_count": len(aligned_rows),
                "columns": sorted({key for row in aligned_rows for key in row.keys()}),
                "application_errors": application_errors,
                "fallback_info": fallback_info,
            }
            self.logger.info("Applied month-aligned card collection for card_id=%s path=%s", dataset.card_id, fallback_path)
            raw_payload = dict(dataset.raw_payload)
            raw_payload["fallback_info"] = fallback_info
            return replace(dataset, rows=aligned_rows, summary=summary, raw_payload=raw_payload)

        if dataset.rows and not application_errors:
            return dataset

        fallback_info = {
            "source_type": "card_collection",
            "fallback_path": str(fallback_path),
            "sheet_name": worksheet.title,
            "replaced_application_errors": bool(application_errors),
            "replaced_empty_result": not dataset.rows and not application_errors,
        }
        summary = {
            "row_count": len(rows),
            "columns": sorted({key for row in rows for key in row.keys()}),
            "application_errors": application_errors,
            "fallback_info": fallback_info,
        }
        self.logger.info("Applied card collection fallback for card_id=%s path=%s", dataset.card_id, fallback_path)
        raw_payload = dict(dataset.raw_payload)
        raw_payload["fallback_info"] = fallback_info
        return replace(dataset, rows=rows, summary=summary, raw_payload=raw_payload)

    def _month_aligned_rows(
        self,
        dataset: NormalizedDataset,
        rows: list[dict[str, Any]],
        report_month: str | None,
    ) -> list[dict[str, Any]] | None:
        if not report_month or dataset.card_id not in self.MONTH_ALIGNED_CARD_IDS:
            return None
        if not self._collection_matches_report_month(report_month):
            return None
        if dataset.role == "order_ratio_anomalies":
            current_month_rows = self._filter_rows_to_report_month(rows, report_month)
            return current_month_rows or None
        return rows

    def _collection_matches_report_month(self, report_month: str) -> bool:
        match = re.search(r"(\d{2})(\d{2})$", self.collection_dir.name)
        if match is None:
            return False
        year_suffix, month = match.groups()
        return report_month == f"20{year_suffix}-{month}"

    def _filter_rows_to_report_month(self, rows: list[dict[str, Any]], report_month: str) -> list[dict[str, Any]]:
        filtered = []
        for row in rows:
            for key in self.PERIOD_HEADERS:
                value = row.get(key)
                if value is not None and str(value).strip().replace("/", "-").startswith(report_month):
                    filtered.append(row)
                    break
        return filtered

    def _resolve_sheet(self, workbook: Any, card_id: str) -> Any | None:
        preferred = self.sheet_overrides.get(card_id)
        if preferred and preferred in workbook.sheetnames:
            return workbook[preferred]
        if workbook.sheetnames:
            return workbook[workbook.sheetnames[0]]
        return None

    def _extract_rows(self, worksheet: Any) -> list[dict[str, Any]]:
        values = list(worksheet.iter_rows(values_only=True))
        if not values:
            return []

        header_row_index = 0
        headers = self._build_headers(values[0], values[1] if len(values) > 1 else None)
        if self._should_use_two_row_header(values[0], values[1] if len(values) > 1 else None):
            header_row_index = 1

        rows: list[dict[str, Any]] = []
        carry_forward_values: dict[str, Any] = {}
        for row_values in values[header_row_index + 1 :]:
            if not row_values or all(value in (None, "") for value in row_values):
                continue
            row: dict[str, Any] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = row_values[index] if index < len(row_values) else None
                normalized_value = self._normalize_value(value)
                if normalized_value is None and header in self.CARRY_FORWARD_HEADERS:
                    normalized_value = carry_forward_values.get(header)
                elif normalized_value is not None and header in self.CARRY_FORWARD_HEADERS:
                    carry_forward_values[header] = normalized_value
                if normalized_value is None:
                    continue
                row[header] = normalized_value
            if row:
                rows.append(row)
        return rows

    def _should_use_two_row_header(self, first_row: tuple[Any, ...], second_row: tuple[Any, ...] | None) -> bool:
        if second_row is None:
            return False
        first = [self._normalize_header(item) for item in first_row]
        second = [self._normalize_header(item) for item in second_row]
        first_nonempty = [item for item in first if item]
        second_nonempty = [item for item in second if item]
        likely_two_row = (
            any(not item for item in first[:2])
            or len(set(first_nonempty)) < len(first_nonempty)
            or (len(second_nonempty) > len(first_nonempty) and len(second_nonempty) >= 3)
        )
        if not likely_two_row:
            return False
        return not self._row_looks_like_data(second_row)

    def _build_headers(self, first_row: tuple[Any, ...], second_row: tuple[Any, ...] | None) -> list[str]:
        first = [self._normalize_header(item) for item in first_row]
        if not self._should_use_two_row_header(first_row, second_row):
            return self._deduplicate_headers(first)

        second = [self._normalize_header(item) for item in second_row or ()]
        merged: list[str] = []
        max_len = max(len(first), len(second))
        for index in range(max_len):
            top = first[index] if index < len(first) else ""
            bottom = second[index] if index < len(second) else ""
            if top and bottom and top != bottom:
                merged.append(f"{top}-{bottom}")
            else:
                merged.append(bottom or top)
        return self._deduplicate_headers(merged)

    def _deduplicate_headers(self, headers: list[str]) -> list[str]:
        counts: dict[str, int] = {}
        result: list[str] = []
        for header in headers:
            base = header.strip()
            if not base:
                result.append("")
                continue
            counts[base] = counts.get(base, 0) + 1
            if counts[base] == 1:
                result.append(base)
            else:
                result.append(f"{base}__{counts[base]}")
        return result

    def _normalize_header(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _row_looks_like_data(self, row: tuple[Any, ...]) -> bool:
        nonempty = [value for value in row if value not in (None, "")]
        if not nonempty:
            return False
        data_like = 0
        for value in nonempty:
            if isinstance(value, (int, float, datetime, date)):
                data_like += 1
                continue
            if isinstance(value, str):
                stripped = value.strip().replace(",", "")
                if not stripped:
                    continue
                if re.fullmatch(r"\d{4}-\d{2}(-\d{2})?", stripped):
                    data_like += 1
                    continue
                try:
                    float(stripped.rstrip("%"))
                    data_like += 1
                    continue
                except ValueError:
                    pass
        return data_like >= max(2, len(nonempty) // 2)

    def _normalize_rows(
        self,
        card_id: str,
        rows: list[dict[str, Any]],
        report_month: str | None = None,
    ) -> list[dict[str, Any]]:
        normalized_rows = rows
        if card_id == "j21833508e589464c922d381":
            normalized_rows = self._normalize_j218_rows(normalized_rows)
        elif card_id == "l1d70dacd48c3422d9f7f67c":
            normalized_rows = self._normalize_l1d70_rows(normalized_rows, report_month=report_month)
        elif card_id == "a597c4441b7414c93a7c502d":
            normalized_rows = self._drop_total_category_rows(normalized_rows, "滔搏纸袋分类")
        elif card_id == "nb692ce19d26a49569de3ca8":
            normalized_rows = self._normalize_nb69_rows(normalized_rows, report_month=report_month)
        return normalized_rows

    def _normalize_j218_rows(self, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        normalized: list[dict[str, Any]] = []
        for row in rows:
            new_row = dict(row)
            first_sync = new_row.pop("同期", None)
            first_yoy = new_row.pop("同比", None)
            second_sync = new_row.pop("同期__2", None)
            second_yoy = new_row.pop("同比__2", None)
            if first_sync is not None:
                new_row["纸袋配比-同期"] = first_sync
            if first_yoy is not None:
                new_row["纸袋配比-同比"] = first_yoy
            if second_sync is not None:
                new_row["同期"] = second_sync
            if second_yoy is not None:
                new_row["同比"] = second_yoy
            normalized.append(new_row)
        return normalized

    def _normalize_l1d70_rows(
        self,
        rows: list[dict[str, Any]],
        report_month: str | None = None,
    ) -> list[dict[str, Any]]:
        normalized: list[dict[str, Any]] = []
        for row in rows:
            paper_bag_type = row.get("纸袋分类")
            bag_model = row.get("滔搏纸袋分类")
            if paper_bag_type is None and isinstance(bag_model, str) and bag_model.startswith("滔搏纸袋-"):
                paper_bag_type = "滔搏纸袋"
            new_row: dict[str, Any] = {}
            if paper_bag_type is not None:
                new_row["纸袋分类"] = paper_bag_type
            new_row["滔搏纸袋分类"] = bag_model or "总计"
            price_values = self._extract_l1d70_price_values(row, report_month=report_month)
            if price_values.get("current") is not None:
                new_row["2026" if not report_month else report_month.split("-")[0]] = price_values["current"]
            if price_values.get("previous") is not None:
                previous_year = "2025"
                if report_month:
                    previous_year = str(int(report_month.split("-")[0]) - 1)
                new_row[previous_year] = price_values["previous"]
            if new_row:
                normalized.append(new_row)
        return normalized

    def _extract_l1d70_price_values(self, row: dict[str, Any], report_month: str | None = None) -> dict[str, Any]:
        metric_keys = [key for key in row.keys() if "纸袋平均单价" in str(key)]
        current_value = None
        previous_value = None
        if report_month:
            current_year = report_month.split("-")[0]
            previous_year = str(int(current_year) - 1)
            for key in metric_keys:
                key_text = str(key)
                if current_year in key_text and current_value is None:
                    current_value = row.get(key)
                elif previous_year in key_text and previous_value is None:
                    previous_value = row.get(key)
            unnamed_values = [row.get(key) for key in metric_keys if current_year not in str(key) and previous_year not in str(key)]
            if current_value is None and unnamed_values:
                current_value = unnamed_values[0]
            if previous_value is None and len(unnamed_values) >= 2:
                previous_value = unnamed_values[1]
        else:
            values = [row.get(key) for key in metric_keys]
            if values:
                current_value = values[0]
            if len(values) >= 2:
                previous_value = values[1]
        return {"current": current_value, "previous": previous_value}

    def _normalize_nb69_rows(self, rows: list[dict[str, Any]], report_month: str | None = None) -> list[dict[str, Any]]:
        if not report_month:
            return rows
        report_year = int(report_month.split("-")[0])
        valid_years = {str(report_year), str(report_year - 1), str(report_year - 2)}
        normalized: list[dict[str, Any]] = []
        for row in rows:
            fiscal_year = row.get("财年")
            if fiscal_year is None:
                continue
            if str(fiscal_year) not in valid_years:
                continue
            normalized.append(row)
        return normalized

    def _drop_total_category_rows(self, rows: list[dict[str, Any]], field: str) -> list[dict[str, Any]]:
        normalized: list[dict[str, Any]] = []
        for row in rows:
            value = row.get(field)
            if isinstance(value, str) and any(token in value for token in self.TOTAL_TOKENS):
                continue
            normalized.append(row)
        return normalized

    def _normalize_value(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, str):
            stripped = value.strip().replace(",", "")
            if not stripped:
                return None
            if stripped.endswith("%"):
                try:
                    return float(stripped[:-1]) / 100
                except ValueError:
                    return stripped
            try:
                if "." in stripped:
                    return float(stripped)
                return int(stripped)
            except ValueError:
                return stripped
        return value
