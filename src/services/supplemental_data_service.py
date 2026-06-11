from __future__ import annotations

import glob
from pathlib import Path
import re
from typing import Any

from src.models.schemas import NormalizedDataset

try:
    import openpyxl
except ImportError:  # pragma: no cover - dependency is declared in requirements.txt
    openpyxl = None


class SupplementalDataService:
    FORECAST_CARD_ID = "u114a0c72ae524037a53c8d1"
    FORECAST_ROLE = "purchase_forecast_sheet"
    FORECAST_SECTION = "inventory_diagnosis"
    FORECAST_NAME = "未来30天纸袋使用量预测-导出表"
    SPECS_CARD_ID = "local_paper_bag_specs_sheet"
    SPECS_ROLE = "paper_bag_specs_reference"
    SPECS_SECTION = "inventory_diagnosis"
    SPECS_NAME = "纸袋规格说明-导出表"

    def __init__(
        self,
        workbook_glob: str,
        sheet_name: str,
        logger: Any,
        workbook_path: str = "",
        paper_bag_specs_workbook_path: str = "",
        paper_bag_specs_sheet_name: str = "纸袋规格",
    ) -> None:
        self.workbook_path = workbook_path
        self.workbook_glob = workbook_glob
        self.sheet_name = sheet_name
        self.paper_bag_specs_workbook_path = paper_bag_specs_workbook_path
        self.paper_bag_specs_sheet_name = paper_bag_specs_sheet_name
        self.logger = logger

    def load_datasets(self, report_month: str | None = None) -> list[NormalizedDataset]:
        datasets: list[NormalizedDataset] = []
        forecast_dataset = self._load_purchase_forecast_sheet(report_month)
        if forecast_dataset is not None:
            datasets.append(forecast_dataset)
        specs_dataset = self._load_paper_bag_specs_sheet()
        if specs_dataset is not None:
            datasets.append(specs_dataset)
        return datasets

    def _load_purchase_forecast_sheet(self, report_month: str | None = None) -> NormalizedDataset | None:
        if openpyxl is None:
            self.logger.warning("openpyxl is not installed, skip loading local forecast workbook.")
            return None

        workbook_path = self._resolve_workbook(report_month)
        if workbook_path is None:
            self.logger.info("No local forecast workbook matched current selection.")
            return None

        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        worksheet = self._resolve_sheet(workbook)
        if worksheet is None:
            self.logger.warning(
                "Forecast workbook %s does not contain sheet %s or compatible fallback sheet.",
                workbook_path,
                self.sheet_name,
            )
            return None

        rows = self._extract_rows(worksheet)
        self.logger.info(
            "Loaded %s rows from forecast workbook %s sheet=%s",
            len(rows),
            workbook_path,
            worksheet.title,
        )
        return NormalizedDataset(
            role=self.FORECAST_ROLE,
            card_id=self.FORECAST_CARD_ID,
            card_name=f"{self.FORECAST_NAME}（{worksheet.title}）",
            section=self.FORECAST_SECTION,
            rows=rows,
            summary={
                "row_count": len(rows),
                "columns": sorted({key for row in rows for key in row.keys()}),
            },
            raw_payload={
                "source_type": "local_workbook",
                "workbook_path": str(workbook_path),
                "sheet_name": worksheet.title,
            },
        )

    def _load_paper_bag_specs_sheet(self) -> NormalizedDataset | None:
        if openpyxl is None:
            self.logger.warning("openpyxl is not installed, skip loading local paper bag specs workbook.")
            return None
        if not self.paper_bag_specs_workbook_path:
            return None

        workbook_path = Path(self.paper_bag_specs_workbook_path).expanduser()
        if not workbook_path.exists():
            self.logger.warning("Configured paper bag specs workbook path does not exist: %s", workbook_path)
            return None

        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        worksheet = workbook[self.paper_bag_specs_sheet_name] if self.paper_bag_specs_sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        rows = self._extract_specs_rows(worksheet)
        self.logger.info(
            "Loaded %s rows from paper bag specs workbook %s sheet=%s",
            len(rows),
            workbook_path,
            worksheet.title,
        )
        return NormalizedDataset(
            role=self.SPECS_ROLE,
            card_id=self.SPECS_CARD_ID,
            card_name=f"{self.SPECS_NAME}（{worksheet.title}）",
            section=self.SPECS_SECTION,
            rows=rows,
            summary={
                "row_count": len(rows),
                "columns": sorted({key for row in rows for key in row.keys()}),
            },
            raw_payload={
                "source_type": "local_workbook",
                "workbook_path": str(workbook_path),
                "sheet_name": worksheet.title,
            },
        )

    def _resolve_workbook(self, report_month: str | None = None) -> Path | None:
        explicit_path = self._resolve_explicit_workbook()
        if explicit_path is not None:
            return explicit_path

        pattern = str(Path(self.workbook_glob).expanduser())
        candidates = sorted((Path(item) for item in glob.glob(pattern)), key=lambda item: item.stat().st_mtime)
        if report_month:
            month_candidates = [item for item in candidates if report_month in item.name]
            if len(month_candidates) == 1:
                return month_candidates[0]
        if candidates:
            if len(candidates) > 1:
                self.logger.warning(
                    "Multiple forecast workbooks matched pattern %s; fallback to latest file %s. Consider setting forecast_workbook_path.",
                    self.workbook_glob,
                    candidates[-1],
                )
            return candidates[-1]
        return None

    def _resolve_explicit_workbook(self) -> Path | None:
        if not self.workbook_path:
            return None
        candidate = Path(self.workbook_path).expanduser()
        if candidate.exists():
            return candidate
        self.logger.warning("Configured forecast workbook path does not exist: %s", candidate)
        return None

    def _resolve_sheet(self, workbook: Any) -> Any | None:
        if self.sheet_name in workbook.sheetnames:
            return workbook[self.sheet_name]
        fallback_names = ["到大区型号", "按大区型号"]
        for name in fallback_names:
            if name in workbook.sheetnames:
                return workbook[name]
        for name in workbook.sheetnames:
            if "大区" in name and "型号" in name:
                return workbook[name]
        return None

    def _extract_rows(self, worksheet: Any) -> list[dict[str, Any]]:
        iterator = worksheet.iter_rows(values_only=True)
        try:
            headers = [self._normalize_header(value) for value in next(iterator)]
        except StopIteration:
            return []

        rows: list[dict[str, Any]] = []
        current_region: str | None = None
        for values in iterator:
            if not values or all(value in (None, "") for value in values):
                continue

            row: dict[str, Any] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = values[index] if index < len(values) else None
                normalized_value = self._normalize_value(value)
                if normalized_value is None:
                    continue
                row[header] = normalized_value

            region = row.get("大区")
            if isinstance(region, str) and region.strip():
                current_region = region.strip()
                row["大区"] = current_region
            elif current_region:
                row["大区"] = current_region

            future_usage = self._find_numeric(row, ["测算未来30天纸袋销量", "同期后30天纸袋销量"])
            current_inventory = self._find_numeric(row, ["筛选日期末库存", "期末库存", "库存"])
            if current_inventory is not None and future_usage not in (None, 0):
                row["次月期末库销测算"] = (current_inventory - future_usage) / future_usage

            if row:
                rows.append(row)
        return rows

    def _extract_specs_rows(self, worksheet: Any) -> list[dict[str, Any]]:
        iterator = worksheet.iter_rows(values_only=True)
        try:
            headers = [self._normalize_header(value) for value in next(iterator)]
        except StopIteration:
            return []

        rows: list[dict[str, Any]] = []
        for values in iterator:
            if not values or all(value in (None, "") for value in values):
                continue

            row: dict[str, Any] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = values[index] if index < len(values) else None
                normalized_value = self._normalize_value(value)
                if normalized_value is None:
                    continue
                row[header] = normalized_value

            code = str(row.get("纸袋编码", "")).strip()
            display_name = str(row.get("规格型号", "")).strip()
            if not code or not display_name or "箱" in display_name:
                continue

            bag_model = self._derive_bag_model_label(code, display_name)
            usage_scene = str(row.get("使用场景", "")).strip().replace("\n", "；")
            row["纸袋型号"] = bag_model
            row["使用场景"] = usage_scene
            rows.append(row)

        return rows

    def _normalize_header(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _normalize_value(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, str):
            stripped = value.strip().replace(",", "")
            if not stripped:
                return None
            if stripped.endswith("%"):
                try:
                    return float(stripped[:-1]) / 100
                except ValueError:
                    return stripped
            try:
                if "." in stripped:
                    return float(stripped)
                return int(stripped)
            except ValueError:
                return stripped
        return value

    def _find_numeric(self, row: dict[str, Any], keys: list[str]) -> float | None:
        matches: list[tuple[int, int, int, float]] = []
        for candidate_position, (candidate_key, value) in enumerate(row.items()):
            if not isinstance(value, (int, float)) or isinstance(value, bool):
                continue
            best_match: tuple[int, int] | None = None
            for index, key in enumerate(keys):
                score = self._match_score(str(candidate_key), key)
                if score > 0:
                    candidate_match = (score, -index)
                    if best_match is None or candidate_match > best_match:
                        best_match = candidate_match
            if best_match is not None:
                matches.append((best_match[0], best_match[1], -candidate_position, float(value)))
        if matches:
            matches.sort(reverse=True)
            return matches[0][3]
        return None

    def _match_score(self, candidate_key: str, key: str) -> int:
        if candidate_key == key:
            return 100
        if candidate_key.startswith(key) or candidate_key.endswith(key):
            return 90
        if key in candidate_key:
            return 80
        return 0

    def _derive_bag_model_label(self, code: str, display_name: str) -> str:
        text = f"{code} {display_name}".upper()
        match = re.search(r"\b(XXS|XS|XL|XXL|S|M|L)\b", text)
        if match:
            return f"滔搏纸袋-{match.group(1)}"
        compact_match = re.search(r"ZD\d+(XXS|XS|XL|XXL|S|M|L)", text)
        if compact_match:
            return f"滔搏纸袋-{compact_match.group(1)}"
        return str(display_name)
