from pathlib import Path
from datetime import datetime

from openpyxl import Workbook

from src.models.schemas import NormalizedDataset
from src.services.card_collection_fallback_service import CardCollectionFallbackService


class DummyLogger:
    def info(self, *_args, **_kwargs) -> None:
        return None


def test_card_collection_fallback_replaces_error_dataset(tmp_path: Path) -> None:
    workbook_path = tmp_path / "demo.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["地区", "订单量"])
    worksheet.append(["东北", 12])
    workbook.save(workbook_path)

    service = CardCollectionFallbackService(tmp_path, DummyLogger())
    datasets = [
        NormalizedDataset(
            role="demo",
            card_id="demo",
            card_name="示例卡",
            section="demo",
            rows=[],
            summary={"row_count": 0, "columns": [], "application_errors": [{"code": 500, "msg": "boom"}]},
            raw_payload={},
        )
    ]

    applied = service.apply(datasets)

    assert applied[0].rows[0]["地区"] == "东北"
    assert applied[0].summary["fallback_info"]["source_type"] == "card_collection"


def test_card_collection_fallback_handles_two_row_headers(tmp_path: Path) -> None:
    workbook_path = tmp_path / "multi.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append([None, None, "2026", "2026"])
    worksheet.append(["纸袋分类", "滔搏纸袋分类", "纸袋销售量", "纸袋平均单价"])
    worksheet.append(["总计", "总计", 100, 1.2])
    workbook.save(workbook_path)

    service = CardCollectionFallbackService(tmp_path, DummyLogger())
    datasets = [
        NormalizedDataset(
            role="demo",
            card_id="multi",
            card_name="示例卡",
            section="demo",
            rows=[],
            summary={"row_count": 0, "columns": [], "application_errors": [{"code": 500, "msg": "boom"}]},
            raw_payload={},
        )
    ]

    applied = service.apply(datasets)

    assert applied[0].rows[0]["纸袋分类"] == "总计"
    assert applied[0].rows[0]["2026-纸袋销售量"] == 100


def test_card_collection_fallback_does_not_treat_duplicate_single_row_header_as_two_row_header(tmp_path: Path) -> None:
    workbook_path = tmp_path / "j21833508e589464c922d381.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["日期 (月)", "纸袋配比", "纸袋配比-不含团购", "同期", "同比", "同期", "同比"])
    worksheet.append(["2026-05", 0.602884, 0.731773, 0.663288, -0.091067, 0.782795, -0.065178])
    workbook.save(workbook_path)

    service = CardCollectionFallbackService(tmp_path, DummyLogger())
    datasets = [
        NormalizedDataset(
            role="consumption_ratio_monthly",
            card_id="j21833508e589464c922d381",
            card_name="配比趋势",
            section="consumption_exceptions",
            rows=[],
            summary={"row_count": 0, "columns": [], "application_errors": [{"code": 500, "msg": "boom"}]},
            raw_payload={},
        )
    ]

    applied = service.apply(datasets)

    assert applied[0].rows == [
        {
            "日期 (月)": "2026-05",
            "纸袋配比": 0.602884,
            "纸袋配比-不含团购": 0.731773,
            "纸袋配比-同期": 0.663288,
            "纸袋配比-同比": -0.091067,
            "同期": 0.782795,
            "同比": -0.065178,
        }
    ]


def test_card_collection_fallback_normalizes_datetime_cells(tmp_path: Path) -> None:
    workbook_path = tmp_path / "trend.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["日期", "FY26-库销比"])
    worksheet.append([datetime(2026, 3, 31), 2.81])
    workbook.save(workbook_path)

    service = CardCollectionFallbackService(tmp_path, DummyLogger())
    datasets = [
        NormalizedDataset(
            role="demo",
            card_id="trend",
            card_name="趋势卡",
            section="demo",
            rows=[],
            summary={"row_count": 0, "columns": [], "application_errors": [{"code": 500, "msg": "boom"}]},
            raw_payload={},
        )
    ]

    applied = service.apply(datasets)

    assert applied[0].rows[0]["日期"] == "2026-03-31"


def test_card_collection_fallback_keeps_api_rows_when_available(tmp_path: Path) -> None:
    workbook_path = tmp_path / "ratio.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["日期 (月)", "纸袋配比", "纸袋配比-不含团购"])
    worksheet.append(["2026-03", 0.61, 0.72])
    workbook.save(workbook_path)

    service = CardCollectionFallbackService(tmp_path, DummyLogger())
    datasets = [
        NormalizedDataset(
            role="consumption_ratio_monthly",
            card_id="ratio",
            card_name="配比趋势",
            section="consumption_exceptions",
            rows=[{"日期 (月)": "2026-03", "纸袋配比": 0.61}],
            summary={"row_count": 1, "columns": ["日期 (月)", "纸袋配比"], "application_errors": []},
            raw_payload={},
        )
    ]

    applied = service.apply(datasets)

    assert applied[0].rows[0] == {"日期 (月)": "2026-03", "纸袋配比": 0.61}
    assert "fallback_info" not in applied[0].summary


def test_card_collection_fallback_carries_forward_model_headers(tmp_path: Path) -> None:
    workbook_path = tmp_path / "model.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["滔搏纸袋分类", "原销售大区", "期末业务库存量"])
    worksheet.append(["滔搏纸袋-L", "小计", 100000])
    worksheet.append([None, "粤海", 38396])
    workbook.save(workbook_path)

    service = CardCollectionFallbackService(tmp_path, DummyLogger())
    datasets = [
        NormalizedDataset(
            role="regional_model_purchase_analysis",
            card_id="model",
            card_name="分型号库存",
            section="inventory_diagnosis",
            rows=[],
            summary={"row_count": 0, "columns": [], "application_errors": [{"code": 500, "msg": "boom"}]},
            raw_payload={},
        )
    ]

    applied = service.apply(datasets)

    assert applied[0].rows[1]["滔搏纸袋分类"] == "滔搏纸袋-L"
    assert applied[0].rows[1]["原销售大区"] == "粤海"


def test_card_collection_fallback_does_not_replace_regional_model_api_rows(tmp_path: Path) -> None:
    workbook_path = tmp_path / "a597.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["滔搏纸袋分类", "原销售大区", "期末近30天累计纸袋销售量", "期末业务库存量", "期末库销比", "期末近30天累计厂入量", "进销比"])
    worksheet.append(["滔搏纸袋-L", "京津晋", 5333, 20214, 3.790362, 15200, 2.850178])
    workbook.save(workbook_path)

    service = CardCollectionFallbackService(tmp_path, DummyLogger())
    datasets = [
        NormalizedDataset(
            role="regional_model_purchase_analysis",
            card_id="a597",
            card_name="分型号库存",
            section="inventory_diagnosis",
            rows=[{"滔搏纸袋分类": "滔搏纸袋-L", "原销售大区": "京津晋", "期末近30天累计纸袋销售量": 6628}],
            summary={"row_count": 1, "columns": ["滔搏纸袋分类", "原销售大区", "期末近30天累计纸袋销售量"], "application_errors": []},
            raw_payload={},
        )
    ]

    applied = service.apply(datasets)

    assert applied[0].rows[0]["期末近30天累计纸袋销售量"] == 6628
    assert "fallback_info" not in applied[0].summary


def test_card_collection_fallback_aligns_known_monthly_cards_for_matching_collection(tmp_path: Path) -> None:
    collection_dir = tmp_path / "纸袋卡片集合-2604"
    collection_dir.mkdir()
    card_id = "xe5da9d423db44bbe96028ad"
    workbook_path = collection_dir / f"{card_id}.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["原销售大区", "期末近30天累计纸袋销售量", "期末业务库存量", "期末库销比"])
    worksheet.append(["川藏新", 72352, 206606, 2.855567])
    workbook.save(workbook_path)

    service = CardCollectionFallbackService(collection_dir, DummyLogger())
    datasets = [
        NormalizedDataset(
            role="regional_inventory_ratio",
            card_id=card_id,
            card_name="大区库销比",
            section="inventory_diagnosis",
            rows=[{"原销售大区": "川藏新", "期末近30天累计纸袋销售量": 88736}],
            summary={"row_count": 1, "columns": ["原销售大区", "期末近30天累计纸袋销售量"], "application_errors": []},
            raw_payload={},
        )
    ]

    applied = service.apply(datasets, report_month="2026-04")

    assert applied[0].rows[0]["期末近30天累计纸袋销售量"] == 72352
    assert applied[0].summary["fallback_info"]["source_type"] == "card_collection_month_alignment"


def test_card_collection_fallback_filters_order_alignment_to_report_month(tmp_path: Path) -> None:
    collection_dir = tmp_path / "纸袋卡片集合-2604"
    collection_dir.mkdir()
    card_id = "l6e08fdcc7fef45ccaa31d1b"
    workbook_path = collection_dir / f"{card_id}.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["日期 (月)", "原销售大区", "原销售店号", "订单编号", "纸袋配比"])
    worksheet.append(["2025-04", "鲁西苏", "A001", "OLD", 10])
    worksheet.append(["2026-04", "鲁东沪", "MTZ068", "MTZ0682604240001", 2.230769])
    workbook.save(workbook_path)

    service = CardCollectionFallbackService(collection_dir, DummyLogger())
    datasets = [
        NormalizedDataset(
            role="order_ratio_anomalies",
            card_id=card_id,
            card_name="订单明细",
            section="consumption_exceptions",
            rows=[{"日期 (月)": "2025-04", "订单编号": "OLD"}],
            summary={"row_count": 1, "columns": ["日期 (月)", "订单编号"], "application_errors": []},
            raw_payload={},
        )
    ]

    applied = service.apply(datasets, report_month="2026-04")

    assert applied[0].rows == [
        {"日期 (月)": "2026-04", "原销售大区": "鲁东沪", "原销售店号": "MTZ068", "订单编号": "MTZ0682604240001", "纸袋配比": 2.230769}
    ]


def test_card_collection_fallback_normalizes_overall_consumption_summary_to_api_shape(tmp_path: Path) -> None:
    workbook_path = tmp_path / "l1d70dacd48c3422d9f7f67c.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append([None, None, "2026", None, None, "2025", None, None])
    worksheet.append(["纸袋分类", "滔搏纸袋分类", "门店纸袋发生费用", "纸袋销售量", "纸袋平均单价(含税)", "门店纸袋发生费用", "纸袋销售量", "纸袋平均单价(含税)"])
    worksheet.append(["总计", None, 3993800.79, 4140825, 0.964494, 2341709.03, 1996342, 1.173])
    worksheet.append(["滔搏纸袋", "小计", 3710708.61, 3963142, 0.936305, 2196716.63, 1908341, 1.151113])
    worksheet.append([None, "滔搏纸袋-XS", 633050.59, 842230, 0.751636, 464892.92, 507351, 0.916314])
    worksheet.append(["非滔搏纸袋", "小计", 283092.18, 177683, 1.593243, 144992.4, 88001, 1.647622])
    worksheet.append([None, "其他", 283092.18, 177683, 1.593243, 144992.4, 88001, 1.647622])
    workbook.save(workbook_path)

    service = CardCollectionFallbackService(tmp_path, DummyLogger())
    datasets = [
        NormalizedDataset(
            role="overall_consumption_summary",
            card_id="l1d70dacd48c3422d9f7f67c",
            card_name="使用情况",
            section="consumption_exceptions",
            rows=[],
            summary={"row_count": 0, "columns": [], "application_errors": [{"code": 500, "msg": "boom"}]},
            raw_payload={},
        )
    ]

    applied = service.apply(datasets)

    assert applied[0].rows == [
        {"纸袋分类": "总计", "滔搏纸袋分类": "总计", "2026": 0.964494, "2025": 1.173},
        {"纸袋分类": "滔搏纸袋", "滔搏纸袋分类": "小计", "2026": 0.936305, "2025": 1.151113},
        {"纸袋分类": "滔搏纸袋", "滔搏纸袋分类": "滔搏纸袋-XS", "2026": 0.751636, "2025": 0.916314},
        {"纸袋分类": "非滔搏纸袋", "滔搏纸袋分类": "小计", "2026": 1.593243, "2025": 1.647622},
        {"纸袋分类": "非滔搏纸袋", "滔搏纸袋分类": "其他", "2026": 1.593243, "2025": 1.647622},
    ]


def test_card_collection_fallback_drops_invalid_stocktake_difference_fiscal_year_rows(tmp_path: Path) -> None:
    workbook_path = tmp_path / "nb692ce19d26a49569de3ca8.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["财年", "盘亏数量", "盘亏金额", "盘盈数量", "盘盈金额"])
    worksheet.append(["2323", 0, 0, 0, 0])
    worksheet.append(["2026", -10, -20, 5, 10])
    worksheet.append(["2025", -30, -40, 15, 20])
    workbook.save(workbook_path)

    service = CardCollectionFallbackService(tmp_path, DummyLogger())
    datasets = [
        NormalizedDataset(
            role="stocktake_difference_chart",
            card_id="nb692ce19d26a49569de3ca8",
            card_name="盘差图表",
            section="consumption_exceptions",
            rows=[],
            summary={"row_count": 0, "columns": [], "application_errors": [{"code": 500, "msg": "boom"}]},
            raw_payload={},
        )
    ]

    applied = service.apply(datasets, report_month="2026-05")

    assert applied[0].rows == [
        {"财年": 2026, "盘亏数量": -10, "盘亏金额": -20, "盘盈数量": 5, "盘盈金额": 10},
        {"财年": 2025, "盘亏数量": -30, "盘亏金额": -40, "盘盈数量": 15, "盘盈金额": 20},
    ]
