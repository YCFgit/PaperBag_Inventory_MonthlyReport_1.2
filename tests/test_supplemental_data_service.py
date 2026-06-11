from pathlib import Path

from openpyxl import Workbook

from src.services.supplemental_data_service import SupplementalDataService


class DummyLogger:
    def info(self, *_args, **_kwargs) -> None:
        return None

    def warning(self, *_args, **_kwargs) -> None:
        return None


def test_supplemental_data_service_loads_sheet_and_fills_region(tmp_path: Path) -> None:
    workbook_path = tmp_path / "滔搏纸袋订购辅助-未来30天纸袋使用量预测 2026-04-08 11_59_34.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "到大区型号"
    worksheet.append(
        [
            "大区",
            "滔搏纸袋分类",
            "筛选日期末库存",
            "近30天纸袋销量",
            "库销比",
            "同期后30天纸袋销量",
            "同期后30天纸袋销量占比",
            "测算未来30天纸袋销量",
            "预估冗余库存量",
            "库存/同期后30天纸袋销量",
        ]
    )
    worksheet.append(["东北", "滔搏纸袋-XS", 52593, 13249, 3.96, 27127, 0.2, 23240.21, 6112.57, 1.94])
    worksheet.append([None, "滔搏纸袋-M", 169206, 94211, 1.8, 104568, 0.79, 89585.38, -9964.77, 1.62])
    workbook.save(workbook_path)

    service = SupplementalDataService(str(workbook_path), "到大区型号", DummyLogger())
    datasets = service.load_datasets()

    assert len(datasets) == 1
    assert datasets[0].card_id == "u114a0c72ae524037a53c8d1"
    assert datasets[0].rows[1]["大区"] == "东北"
    assert round(datasets[0].rows[0]["次月期末库销测算"], 2) == 1.26


def test_supplemental_data_service_prefers_explicit_workbook_path(tmp_path: Path) -> None:
    explicit_path = tmp_path / "explicit.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "到大区型号"
    worksheet.append(["大区", "滔搏纸袋分类", "筛选日期末库存", "测算未来30天纸袋销量"])
    worksheet.append(["东北", "滔搏纸袋-XS", 100, 50])
    workbook.save(explicit_path)

    service = SupplementalDataService(
        workbook_glob=str(tmp_path / "*.xlsx"),
        sheet_name="到大区型号",
        logger=DummyLogger(),
        workbook_path=str(explicit_path),
    )
    datasets = service.load_datasets(report_month="2026-03")

    assert len(datasets) == 1
    assert datasets[0].raw_payload["workbook_path"] == str(explicit_path)


def test_supplemental_data_service_loads_paper_bag_specs_reference(tmp_path: Path) -> None:
    specs_path = tmp_path / "纸袋规格.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "纸袋规格"
    worksheet.append(["纸袋编码", "规格型号", "型号(mm)", "使用场景", "使用频率"])
    worksheet.append(["ZD2023XL", "1号（特大号纸袋XL）", "ZD2023XL(55X42+15)", "棉羽\n特殊鞋盒", "冬季订购"])
    worksheet.append([None, "100/箱", None, None, None])
    worksheet.append(["ZD2023M", "3号(中号纸袋M)", "ZD2023M(29X42+13)", "常规鞋盒一双", "最常用"])
    workbook.save(specs_path)

    service = SupplementalDataService(
        workbook_glob=str(tmp_path / "*.xlsx"),
        sheet_name="到大区型号",
        logger=DummyLogger(),
        paper_bag_specs_workbook_path=str(specs_path),
        paper_bag_specs_sheet_name="纸袋规格",
    )
    datasets = service.load_datasets(report_month="2026-03")

    assert len(datasets) == 1
    assert datasets[0].role == "paper_bag_specs_reference"
    assert datasets[0].rows[0]["纸袋型号"] == "滔搏纸袋-XL"
    assert datasets[0].rows[0]["使用场景"] == "棉羽；特殊鞋盒"
